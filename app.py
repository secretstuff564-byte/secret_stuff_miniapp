from flask import Flask, render_template, abort, redirect, url_for, request, jsonify
import pandas as pd
import json
from openpyxl import load_workbook
import requests  # добавлен импорт

app = Flask(__name__)

# ===== НАСТРОЙКИ ТЕЛЕГРАМ =====
TELEGRAM_TOKEN = '8262290622:AAF5Zp0hgJr5fOJ4nn6JhOd4CFV0GJNk8Bg'
TELEGRAM_CHAT_ID = '@secret_manager_1'

def send_telegram_message(text):
    """Отправляет сообщение менеджеру через Telegram Bot API"""
    url = f'https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage'
    payload = {
        'chat_id': TELEGRAM_CHAT_ID,
        'text': text,
        'parse_mode': 'HTML'
    }
    try:
        response = requests.post(url, json=payload)
        if not response.ok:
            print(f'Ошибка отправки в Telegram: {response.text}')
        return response.ok
    except Exception as e:
        print(f'Ошибка отправки в Telegram: {e}')
        return False


# ===== ФИЛЬТР ДЛЯ ФОРМАТИРОВАНИЯ ЦЕН =====
@app.template_filter('to_locale')
def to_locale(value):
    try:
        return f"{int(value):,}".replace(',', ' ')
    except (ValueError, TypeError):
        return value


EXCEL_PATH = 'products.xlsx'


def load_products():
    df = pd.read_excel(EXCEL_PATH)
    df = df.dropna(subset=['id'])
    df['id'] = df['id'].astype(int)
    products = df.to_dict(orient='records')

    for p in products:
        # --- Обработка images ---
        if isinstance(p['images'], str):
            p['images'] = [img.strip() for img in p['images'].split(',') if img.strip()]
        else:
            p['images'] = []

        # --- lookbook_images ---
        lookbook_str = p.get('lookbook_images', '')
        if isinstance(lookbook_str, str) and lookbook_str.strip():
            p['lookbook_images'] = [img.strip() for img in lookbook_str.split(',') if img.strip()]
        else:
            p['lookbook_images'] = []

        # --- variants ---
        variants_str = p.get('variants', '{}')
        if isinstance(variants_str, str) and variants_str.strip():
            try:
                p['variants'] = json.loads(variants_str)
            except:
                p['variants'] = {}
        else:
            p['variants'] = {}

        # Для обратной совместимости
        if not p['variants'] and 'sizes_stock' in p and pd.notna(p['sizes_stock']):
            old_sizes = {}
            sizes_str = p['sizes_stock']
            if isinstance(sizes_str, str):
                for part in sizes_str.split(','):
                    part = part.strip()
                    if ':' in part:
                        size, stock = part.split(':', 1)
                        try:
                            old_sizes[size.strip()] = int(stock.strip())
                        except:
                            pass
            if old_sizes:
                p['variants'] = {"": old_sizes}

        # --- has_non_one_size ---
        if p.get('variants'):
            has_non_one_size = any(
                size != 'one_size'
                for sizes in p['variants'].values()
                for size in sizes.keys()
            )
            p['has_non_one_size'] = has_non_one_size
        else:
            p['has_non_one_size'] = False

        # --- description ---
        if 'description' not in p or pd.isna(p['description']):
            p['description'] = ''

        # --- featured ---
        featured_val = p.get('featured')
        if pd.isna(featured_val):
            p['featured'] = 0
        else:
            try:
                p['featured'] = int(featured_val)
            except (ValueError, TypeError):
                p['featured'] = 0

    return products


products = load_products()

CATEGORY_NAMES = {
    'footwear': 'Обувь',
    't-shirts': 'Футболки',
    'jeans': 'Джинсы',
    'hoodies': 'Худи',
    'shorts': 'Шорты',
    'puffer': 'Пуховики',
    'windbreakers': 'Ветровки',
    'bags': 'Сумки',
    'accessories': 'Аксессуары',
    'clothing': 'Одежда'
}


def get_products_by_category(category):
    return [p for p in products if p.get('category', '').lower() == category.lower()]


# ---------- МАРШРУТЫ ----------
@app.route('/')
def index():
    featured_products = [p for p in products if p.get('featured', 0) == 1]
    return render_template('index.html', products=featured_products)


@app.route('/category/<string:cat>')
def category(cat):
    filtered = get_products_by_category(cat)
    category_ru = CATEGORY_NAMES.get(cat, cat.capitalize())
    return render_template('category.html', products=filtered, category=cat, category_ru=category_ru)


@app.route('/clothing')
def clothing():
    return redirect(url_for('category', cat='clothing'))


@app.route('/footwear')
def footwear():
    return redirect(url_for('category', cat='footwear'))


@app.route('/bags')
def bags():
    return redirect(url_for('category', cat='bags'))


@app.route('/accessories')
def accessories():
    return redirect(url_for('category', cat='accessories'))


@app.route('/product/<int:product_id>')
def product_detail(product_id):
    product = next((p for p in products if p['id'] == product_id), None)
    if not product:
        abort(404)
    return render_template('product.html', product=product)


@app.route('/cart')
def cart():
    return render_template('cart.html', products=products)


@app.route('/favorites')
def favorites():
    return render_template('favorites.html', products=products)


@app.route('/reload')
def reload_products():
    global products
    products = load_products()
    return 'Товары перезагружены из Excel'


# ===== МАРШРУТ ДЛЯ ОФОРМЛЕНИЯ ЗАКАЗА (С СОХРАНЕНИЕМ ФОРМАТИРОВАНИЯ И УВЕДОМЛЕНИЕМ) =====
@app.route('/checkout', methods=['POST'])
def checkout():
    data = request.get_json()
    if not data or 'cart' not in data:
        return jsonify({'success': False, 'error': 'Нет данных корзины'}), 400

    cart = data['cart']
    total = data.get('total', 0)

    try:
        # Загружаем книгу Excel с форматированием
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # Находим колонки по заголовкам (первая строка)
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers[cell_value] = col

        if 'id' not in headers or 'variants' not in headers:
            return jsonify({'success': False, 'error': 'Не найдены нужные колонки'}), 400

        id_col = headers['id']
        variants_col = headers['variants']

        # Обрабатываем каждый товар в корзине
        for item in cart:
            product_id = item['id']
            color = item.get('color', '')
            size = item.get('size', '')
            quantity = item['quantity']

            # Ищем строку с нужным id
            row_found = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=id_col).value == product_id:
                    row_found = row
                    break

            if row_found is None:
                return jsonify({'success': False, 'error': f'Товар с id {product_id} не найден'}), 400

            # Читаем текущие варианты
            variants_cell = ws.cell(row=row_found, column=variants_col)
            variants_str = variants_cell.value
            if not variants_str:
                return jsonify({'success': False, 'error': f'У товара {product_id} нет вариантов'}), 400

            try:
                variants = json.loads(variants_str)
            except:
                return jsonify({'success': False, 'error': f'Ошибка в формате variants для товара {product_id}'}), 400

            if color not in variants:
                return jsonify({'success': False, 'error': f'Цвет "{color}" не найден для товара {product_id}'}), 400
            if size not in variants[color]:
                return jsonify({'success': False, 'error': f'Размер "{size}" не найден для цвета "{color}" товара {product_id}'}), 400

            current_stock = variants[color][size]
            if current_stock < quantity:
                return jsonify({'success': False, 'error': f'Недостаточно остатков для товара {product_id}, цвет {color}, размер {size}. Доступно: {current_stock}'}), 400

            # Уменьшаем остаток
            variants[color][size] -= quantity
            # Обновляем ячейку (сохраняя формат)
            variants_cell.value = json.dumps(variants, ensure_ascii=False)

        # Сохраняем книгу — форматирование не сбросится
        wb.save(EXCEL_PATH)

        # Перезагружаем глобальный список товаров
        global products
        products = load_products()

        # ===== ОТПРАВКА УВЕДОМЛЕНИЯ МЕНЕДЖЕРУ =====
        message = f"🛒 <b>Новый заказ!</b>\n\n"
        for item in cart:
            product_name = item.get('name', 'Товар')
            color = item.get('color', '—')
            size = item.get('size', '—')
            qty = item.get('quantity', 1)
            price = item.get('price', 0)
            message += f"• {product_name} (Цвет: {color}, Размер: {size}) × {qty} = {price * qty} ₽\n"
        message += f"\n💰 <b>Итого: {total} ₽</b>"

        send_telegram_message(message)

        return jsonify({'success': True, 'message': 'Заказ оформлен, остатки обновлены'})

    except PermissionError:
        return jsonify({'success': False, 'error': 'Файл Excel занят. Закройте Excel и попробуйте снова.'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': f'Ошибка при обработке заказа: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True)